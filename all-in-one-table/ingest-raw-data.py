#!/usr/bin/env python3
"""
Ingest raw CSV and XLSX data from multiple systems into a unified pandas DataFrame.
Uses dataclasses to represent system data with metadata.
"""

import csv
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook


class FileType(Enum):
    """Supported file types for data ingestion."""

    CSV = "csv"
    XLSX = "xlsx"


@dataclass
class SystemRecord:
    """Represents a single record from a system CSV file."""

    hostname: str
    fqdn: str
    ip_address: str
    cpu_usage: float
    memory_usage: float
    disk_usage: float
    network_in: float
    network_out: float
    process_count: int
    uptime_hours: int


@dataclass
class SystemData:
    """Container for system data with metadata."""

    system_name: str
    date_generated: datetime
    records: List[SystemRecord] = field(default_factory=list)

    @classmethod
    def from_csv(cls, filepath: Path, system_name: str) -> "SystemData":
        """Load system data from a CSV file."""
        records = []
        with open(filepath, "r", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                record = SystemRecord(
                    hostname=row["hostname"],
                    fqdn=row["fqdn"],
                    ip_address=row["ip_address"],
                    cpu_usage=float(row["cpu_usage"]),
                    memory_usage=float(row["memory_usage"]),
                    disk_usage=float(row["disk_usage"]),
                    network_in=float(row["network_in"]),
                    network_out=float(row["network_out"]),
                    process_count=int(row["process_count"]),
                    uptime_hours=int(row["uptime_hours"]),
                )
                records.append(record)

        return cls(
            system_name=system_name,
            date_generated=datetime.now(),
            records=records,
        )

    @classmethod
    def from_xlsx(
        cls, filepath: Path, system_name: str, sheet_name: Optional[str] = None
    ) -> "SystemData":
        """Load system data from an XLSX file.

        Args:
            filepath: Path to the XLSX file
            system_name: Name to identify this system
            sheet_name: Optional sheet name. If None, uses the first sheet.
        """
        records = []
        wb = load_workbook(filepath, read_only=True, data_only=True)

        # Use specified sheet or first sheet
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Get headers from first row
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return cls(
                system_name=system_name, date_generated=datetime.now(), records=[]
            )

        headers = [str(h).lower() if h else "" for h in rows[0]]

        # Process data rows
        for row in rows[1:]:
            if not any(row):  # Skip empty rows
                continue
            row_dict = dict(zip(headers, row))
            record = SystemRecord(
                hostname=str(row_dict.get("hostname", "")),
                fqdn=str(row_dict.get("fqdn", "")),
                ip_address=str(row_dict.get("ip_address", "")),
                cpu_usage=float(row_dict.get("cpu_usage", 0)),
                memory_usage=float(row_dict.get("memory_usage", 0)),
                disk_usage=float(row_dict.get("disk_usage", 0)),
                network_in=float(row_dict.get("network_in", 0)),
                network_out=float(row_dict.get("network_out", 0)),
                process_count=int(row_dict.get("process_count", 0)),
                uptime_hours=int(row_dict.get("uptime_hours", 0)),
            )
            records.append(record)

        wb.close()
        return cls(
            system_name=system_name,
            date_generated=datetime.now(),
            records=records,
        )

    @classmethod
    def from_file(
        cls,
        filepath: Path,
        system_name: str,
        file_type: FileType,
        sheet_name: Optional[str] = None,
    ) -> "SystemData":
        """Load system data from a file based on file type.

        Args:
            filepath: Path to the data file
            system_name: Name to identify this system
            file_type: Type of file (CSV or XLSX)
            sheet_name: Optional sheet name for XLSX files
        """
        if file_type == FileType.CSV:
            return cls.from_csv(filepath, system_name)
        elif file_type == FileType.XLSX:
            return cls.from_xlsx(filepath, system_name, sheet_name)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def to_dataframe(self) -> pd.DataFrame:
        """Convert records to a pandas DataFrame with system_name column."""
        data = []
        for record in self.records:
            row = {
                "system_name": self.system_name,
                "date_generated": self.date_generated,
                "hostname": record.hostname,
                "fqdn": record.fqdn,
                "ip_address": record.ip_address,
                "cpu_usage": record.cpu_usage,
                "memory_usage": record.memory_usage,
                "disk_usage": record.disk_usage,
                "network_in": record.network_in,
                "network_out": record.network_out,
                "process_count": record.process_count,
                "uptime_hours": record.uptime_hours,
            }
            data.append(row)
        return pd.DataFrame(data)


@dataclass
class FileConfig:
    """Configuration for a data file to load."""

    filename: str
    system_name: str
    file_type: FileType
    sheet_name: Optional[str] = None  # For XLSX files


def load_all_systems(data_dir: Path) -> tuple[List[SystemData], pd.DataFrame]:
    """
    Load all data files from the data directory.

    Returns:
        A tuple containing:
        - List of SystemData objects (one per file)
        - Combined pandas DataFrame with all records
    """
    # File configurations with type indicators
    file_configs = [
        FileConfig("system_alpha.csv", "Alpha", FileType.CSV),
        FileConfig("system_beta.csv", "Beta", FileType.CSV),
        FileConfig("system_gamma.csv", "Gamma", FileType.CSV),
        FileConfig(
            "system_delta.xlsx", "Delta", FileType.XLSX, sheet_name="SystemData"
        ),
        FileConfig(
            "system_epsilon.xlsx", "Epsilon", FileType.XLSX, sheet_name="SystemData"
        ),
    ]

    # Store SystemData objects in an array
    system_data_array: List[SystemData] = []

    for config in file_configs:
        filepath = data_dir / config.filename
        if filepath.exists():
            system_data = SystemData.from_file(
                filepath, config.system_name, config.file_type, config.sheet_name
            )
            system_data_array.append(system_data)
            print(
                f"Loaded {len(system_data.records)} records from {config.system_name} ({config.file_type.value})"
            )

    # Combine all data into a single DataFrame
    dataframes = [sd.to_dataframe() for sd in system_data_array]
    combined_df = pd.concat(dataframes, ignore_index=True)

    return system_data_array, combined_df


def main():
    """Main entry point."""
    data_dir = Path(__file__).parent / "data"

    print("Loading system data...")
    print("-" * 50)

    system_data_array, combined_df = load_all_systems(data_dir)

    print("-" * 50)
    print(f"\nTotal systems loaded: {len(system_data_array)}")
    print(f"Total records: {len(combined_df)}")

    print("\n" + "=" * 50)
    print("System Data Array Contents:")
    print("=" * 50)
    for sd in system_data_array:
        print(f"\nSystem: {sd.system_name}")
        print(f"  Date Generated: {sd.date_generated}")
        print(f"  Record Count: {len(sd.records)}")

    print("\n" + "=" * 50)
    print("Combined DataFrame:")
    print("=" * 50)
    print(combined_df.to_string())

    print("\n" + "=" * 50)
    print("DataFrame Info:")
    print("=" * 50)
    print(combined_df.info())

    print("\n" + "=" * 50)
    print("Statistics by System:")
    print("=" * 50)
    print(
        combined_df.groupby("system_name")[
            ["cpu_usage", "memory_usage", "disk_usage"]
        ].mean()
    )


if __name__ == "__main__":
    main()
